"""
Juntador de Planilhas - Comitê de Cadastros
=============================================
Como usar:
  1. Coloque todas as planilhas recebidas dos compradores na pasta de entrada
  2. Execute: python juntar_planilhas.py
  3. Informe a data do comitê quando solicitado
  4. A planilha unificada será gerada automaticamente na pasta de saída
"""

import sys, subprocess, os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Instalando openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--user"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Pastas fixas ──────────────────────────────────────────────────────────────
PASTA_ENTRADA = r"X:\Central de Negocios\Cadastro\Fabian\Automação - Ata Comitê de Cadastro\Base Dados - Compradores"
PASTA_SAIDA   = r"X:\Central de Negocios\Cadastro\Fabian\Automação - Ata Comitê de Cadastro\Base Dados Completa - Geração da Ata"


def juntar_planilhas(pasta_entrada, pasta_saida, data_comite):

    arquivos = [
        f for f in os.listdir(pasta_entrada)
        if f.endswith(".xlsx") and not f.startswith("~")
    ]

    if not arquivos:
        print("Nenhum arquivo .xlsx encontrado na pasta de entrada.")
        return False

    print(f"\n{len(arquivos)} arquivo(s) encontrado(s):")
    for f in arquivos:
        print(f"  - {f}")

    # Lê e junta todos os dados
    todos_dados = []
    for arquivo in arquivos:
        caminho = os.path.join(pasta_entrada, arquivo)
        try:
            wb = openpyxl.load_workbook(caminho)
            ws = wb.active
            linhas_adicionadas = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                comprador    = str(row[0]).strip() if row[0] else ""
                cod_barras   = str(row[1]).strip() if row[1] else ""
                nome_produto = str(row[2]).strip() if row[2] else ""
                if not nome_produto:
                    continue
                todos_dados.append((comprador, cod_barras, nome_produto))
                linhas_adicionadas += 1
            print(f"  ✓ {arquivo}: {linhas_adicionadas} item(ns) lido(s)")
        except Exception as e:
            print(f"  ✗ {arquivo}: erro ao ler — {e}")

    if not todos_dados:
        print("\nNenhum dado encontrado nas planilhas.")
        return False

    # Cria a planilha de saída
    wb_saida = openpyxl.Workbook()
    ws_saida = wb_saida.active
    ws_saida.title = "Itens Reunião"

    # Estilos do cabeçalho
    fonte_header  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    fill_header   = PatternFill("solid", fgColor="1F3864")
    alinha_centro = Alignment(horizontal="center", vertical="center")

    cabecalho = ["Comprador", "Cod. Barras", "Nome do Produto"]
    for col, titulo in enumerate(cabecalho, start=1):
        cel = ws_saida.cell(row=1, column=col, value=titulo)
        cel.font      = fonte_header
        cel.fill      = fill_header
        cel.alignment = alinha_centro

    ws_saida.column_dimensions["A"].width = 22
    ws_saida.column_dimensions["B"].width = 18
    ws_saida.column_dimensions["C"].width = 55
    ws_saida.row_dimensions[1].height    = 22

    # Preenche os dados
    fonte_dados     = Font(name="Arial", size=10)
    fonte_comprador = Font(name="Arial", bold=True, size=10)
    fill_comp       = PatternFill("solid", fgColor="D6E4F0")
    fill_par        = PatternFill("solid", fgColor="EBF3FB")
    fill_impar      = PatternFill("solid", fgColor="FFFFFF")

    comprador_atual = None
    linha = 2

    for comprador, cod, nome in todos_dados:
        # Linha de destaque quando muda o comprador
        if comprador != comprador_atual:
            comprador_atual = comprador
            for col in range(1, 4):
                cel      = ws_saida.cell(row=linha, column=col)
                cel.fill = fill_comp
                cel.font = fonte_comprador
            ws_saida.cell(row=linha, column=1, value=comprador)
            ws_saida.cell(row=linha, column=2, value="")
            ws_saida.cell(row=linha, column=3, value="")
            linha += 1

        fill_item = fill_par if linha % 2 == 0 else fill_impar
        for col, val in enumerate([comprador, cod, nome], start=1):
            cel           = ws_saida.cell(row=linha, column=col, value=val)
            cel.font      = fonte_dados
            cel.fill      = fill_item
            cel.alignment = Alignment(vertical="center")
        linha += 1

    ws_saida.freeze_panes = "A2"

    # Nome do arquivo com a data do comitê
    nome_saida    = f"Itens Reunião de Comitê - {data_comite}.xlsx"
    caminho_saida = os.path.join(pasta_saida, nome_saida)
    wb_saida.save(caminho_saida)

    return caminho_saida, len(todos_dados)


def main():
    print("=" * 50)
    print("  Juntador de Planilhas - Comitê de Cadastros")
    print("=" * 50)

    # Verifica pastas
    if not os.path.isdir(PASTA_ENTRADA):
        print(f"\n❌ Pasta de entrada não encontrada:\n   {PASTA_ENTRADA}")
        sys.exit(1)
    if not os.path.isdir(PASTA_SAIDA):
        try:
            os.makedirs(PASTA_SAIDA)
            print(f"Pasta de saída criada: {PASTA_SAIDA}")
        except Exception as e:
            print(f"\n❌ Não foi possível criar a pasta de saída: {e}")
            sys.exit(1)

    # Input da data do comitê
    while True:
        entrada = input("\nQual a data do comitê? (DD/MM/AAAA) ou Enter para hoje: ").strip()
        if entrada == "":
            hoje       = datetime.today()
            data_comite = f"{hoje.day:02d} {hoje.month:02d} {hoje.year}"
            break
        try:
            data        = datetime.strptime(entrada, "%d/%m/%Y")
            data_comite = f"{data.day:02d} {data.month:02d} {data.year}"
            break
        except ValueError:
            print("  Formato inválido. Use DD/MM/AAAA (ex: 05/06/2026)")

    print("\nProcessando...")
    resultado = juntar_planilhas(PASTA_ENTRADA, PASTA_SAIDA, data_comite)

    if resultado:
        caminho_saida, total = resultado
        print(f"\n✅ Planilha unificada gerada com sucesso!")
        print(f"   Total de itens: {total}")
        print(f"   Arquivo: '{caminho_saida}'")
    else:
        print("\n❌ Falha ao gerar a planilha unificada.")
        sys.exit(1)


if __name__ == "__main__":
    main()
